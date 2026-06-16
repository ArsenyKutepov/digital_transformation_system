import json
import csv
from datetime import datetime
from database.connection import DatabaseConnection


class ReportGenerator:
    """Генерация отчётов в различных форматах"""

    def __init__(self):
        self.db = DatabaseConnection()

    def export_to_excel(self, org_id: int, output_path: str) -> bool:
        """Экспорт в Excel (использует openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.chart import LineChart, Reference

            wb = openpyxl.Workbook()

            # Лист с оценками
            ws_assessments = wb.active
            ws_assessments.title = "Оценки ЦЗ"

            # Заголовки
            headers = ['Дата', 'Интегральный балл', 'Технический фактор', 'Когнитивный фактор', 'Личностный фактор']
            for col, header in enumerate(headers, 1):
                cell = ws_assessments.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            # Данные
            assessments = self.db.query(
                "SELECT * FROM assessments WHERE org_id = ? ORDER BY assessment_date",
                (org_id,)
            )
            for row, a in enumerate(assessments, 2):
                ws_assessments.cell(row=row, column=1, value=a['assessment_date'])
                ws_assessments.cell(row=row, column=2, value=a['final_score'])
                ws_assessments.cell(row=row, column=3, value=a['technical_factor'])
                ws_assessments.cell(row=row, column=4, value=a['cognitive_factor'])
                ws_assessments.cell(row=row, column=5, value=a['personal_factor'])

            # График
            if len(assessments) > 1:
                chart = LineChart()
                chart.title = "Динамика цифровой зрелости"
                chart.style = 13
                chart.y_axis.title = "Баллы"
                chart.x_axis.title = "Дата"

                data = Reference(ws_assessments, min_col=2, min_row=1, max_row=len(assessments) + 1, max_col=2)
                dates = Reference(ws_assessments, min_col=1, min_row=2, max_row=len(assessments) + 1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(dates)
                ws_assessments.add_chart(chart, "G2")

            # Лист с дорожной картой
            plan = self.db.query_one(
                "SELECT * FROM transformation_plans WHERE org_id = ? ORDER BY plan_date DESC LIMIT 1",
                (org_id,)
            )

            if plan:
                ws_plan = wb.create_sheet("Дорожная карта")
                plan_data = json.loads(plan['plan_data']) if plan['plan_data'] else []

                headers_plan = ['Шаг', 'Действие', 'Ожидаемый рост', 'Затраты', 'Риск']
                for col, header in enumerate(headers_plan, 1):
                    cell = ws_plan.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)

                for row, action in enumerate(plan_data, 2):
                    ws_plan.cell(row=row, column=1, value=action.get('step', row - 1))
                    ws_plan.cell(row=row, column=2, value=action.get('action_name', ''))
                    ws_plan.cell(row=row, column=3, value=action.get('base_growth', 0))
                    ws_plan.cell(row=row, column=4, value=action.get('cost', 0))
                    ws_plan.cell(row=row, column=5, value=action.get('risk', 0))

            # Лист со спринтами
            if plan:
                ws_sprints = wb.create_sheet("Спринты")
                sprints = self.db.query(
                    "SELECT * FROM sprints WHERE plan_id = ? ORDER BY number",
                    (plan['id'],)
                )

                headers_sprints = ['Номер', 'Статус', 'Плановый старт', 'Плановый финиш', 'Фактический старт',
                                   'Фактический финиш', 'Плановый рост', 'Фактический рост']
                for col, header in enumerate(headers_sprints, 1):
                    cell = ws_sprints.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)

                for row, s in enumerate(sprints, 2):
                    ws_sprints.cell(row=row, column=1, value=s['number'])
                    ws_sprints.cell(row=row, column=2, value=s['status'])
                    ws_sprints.cell(row=row, column=3, value=s['planned_start'])
                    ws_sprints.cell(row=row, column=4, value=s['planned_end'])
                    ws_sprints.cell(row=row, column=5, value=s['actual_start'] or '')
                    ws_sprints.cell(row=row, column=6, value=s['actual_end'] or '')
                    ws_sprints.cell(row=row, column=7, value=s['planned_growth'])
                    ws_sprints.cell(row=row, column=8, value=s['actual_growth'] or '')

            wb.save(output_path)
            return True

        except ImportError:
            return False

    def export_to_json(self, org_id: int) -> str:
        """Экспорт в JSON"""
        assessments = self.db.query(
            "SELECT * FROM assessments WHERE org_id = ? ORDER BY assessment_date",
            (org_id,)
        )

        plan = self.db.query_one(
            "SELECT * FROM transformation_plans WHERE org_id = ? ORDER BY plan_date DESC LIMIT 1",
            (org_id,)
        )

        completed = self.db.query(
            "SELECT ca.*, a.name FROM completed_actions ca JOIN actions a ON ca.action_id = a.id WHERE ca.org_id = ?",
            (org_id,)
        )

        result = {
            'export_date': datetime.now().isoformat(),
            'org_id': org_id,
            'assessments': [
                {
                    'date': a['assessment_date'],
                    'final_score': a['final_score'],
                    'technical_factor': a['technical_factor'],
                    'cognitive_factor': a['cognitive_factor'],
                    'personal_factor': a['personal_factor']
                }
                for a in assessments
            ],
            'plan': json.loads(plan['plan_data']) if plan and plan['plan_data'] else [],
            'completed_actions': [
                {
                    'action_name': c['name'],
                    'completed_date': c['completed_date'],
                    'actual_growth': c['actual_growth']
                }
                for c in completed
            ]
        }

        return json.dumps(result, ensure_ascii=False, indent=2)

    def export_to_csv(self, org_id: int, output_path: str) -> bool:
        """Экспорт в CSV"""
        assessments = self.db.query(
            "SELECT * FROM assessments WHERE org_id = ? ORDER BY assessment_date",
            (org_id,)
        )

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(
                ['Дата', 'Интегральный балл', 'Технический фактор', 'Когнитивный фактор', 'Личностный фактор'])

            for a in assessments:
                writer.writerow([
                    a['assessment_date'],
                    a['final_score'],
                    a['technical_factor'],
                    a['cognitive_factor'],
                    a['personal_factor']
                ])

        return True

    def generate_dashboard_html(self, org_id: int) -> str:
        """Генерация HTML-дашборда для организации"""
        assessments = self.db.query(
            "SELECT * FROM assessments WHERE org_id = ? ORDER BY assessment_date",
            (org_id,)
        )

        org = self.db.query_one("SELECT name FROM organizations WHERE id = ?", (org_id,))
        org_name = org['name'] if org else 'Организация'

        # Подготовка данных для графиков
        dates = [a['assessment_date'][:10] for a in assessments]
        scores = [a['final_score'] for a in assessments]
        technical = [a['technical_factor'] for a in assessments]
        cognitive = [a['cognitive_factor'] for a in assessments]
        personal = [a['personal_factor'] for a in assessments]

        latest_score = scores[-1] if scores else 0
        first_score = scores[0] if scores else 0
        growth = latest_score - first_score if scores else 0

        html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Дашборд - {org_name}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #0F172A; color: #F1F5F9; }}
        .container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(135deg, #1E293B, #0F172A); padding: 30px; border-radius: 20px; margin-bottom: 20px; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; margin-bottom: 30px; }}
        .stat-card {{ background: #1E293B; border-radius: 16px; padding: 20px; text-align: center; border: 1px solid #334155; }}
        .stat-number {{ font-size: 36px; font-weight: bold; color: #3B82F6; }}
        .stat-label {{ font-size: 14px; color: #94A3B8; margin-top: 8px; }}
        .chart-card {{ background: #1E293B; border-radius: 16px; padding: 20px; margin-bottom: 20px; border: 1px solid #334155; }}
        .chart-card h3 {{ margin-bottom: 15px; color: #F1F5F9; }}
        .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 Дашборд цифровой трансформации</h1>
            <p>{org_name} | Актуально на {datetime.now().strftime('%d.%m.%Y')}</p>
        </div>

        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-number">{latest_score:.1f}</div>
                <div class="stat-label">Текущий уровень ЦЗ</div>
            </div>
            <div class="stat-card">
                <div class="stat-number">{growth:+.1f}</div>
                <div class="stat-label">Рост за период</div>
            </div>
            <div class="stat-card">
                <div class="stat-number">{len(assessments)}</div>
                <div class="stat-label">Выполнено оценок</div>
            </div>
            <div class="stat-card">
                <div class="stat-number">25</div>
                <div class="stat-label">Показателей</div>
            </div>
        </div>

        <div class="chart-card">
            <h3>📈 Тренд цифровой зрелости</h3>
            <canvas id="trendChart" style="height: 300px;"></canvas>
        </div>

        <div class="grid-2">
            <div class="chart-card">
                <h3>📊 Факторы ЦЗ</h3>
                <canvas id="factorsChart" style="height: 250px;"></canvas>
            </div>
            <div class="chart-card">
                <h3>🎯 Динамика факторов</h3>
                <canvas id="factorsTrendChart" style="height: 250px;"></canvas>
            </div>
        </div>
    </div>

    <script>
        const dates = {json.dumps(dates)};
        const scores = {json.dumps(scores)};
        const technical = {json.dumps(technical)};
        const cognitive = {json.dumps(cognitive)};
        const personal = {json.dumps(personal)};

        new Chart(document.getElementById('trendChart'), {{
            type: 'line',
            data: {{
                labels: dates,
                datasets: [{{
                    label: 'Цифровая зрелость (баллы)',
                    data: scores,
                    borderColor: '#3B82F6',
                    backgroundColor: 'rgba(59,130,246,0.1)',
                    fill: true,
                    tension: 0.3
                }}]
            }},
            options: {{ responsive: true, maintainAspectRatio: true }}
        }});

        if (technical.length > 0) {{
            const lastIndex = technical.length - 1;
            new Chart(document.getElementById('factorsChart'), {{
                type: 'bar',
                data: {{
                    labels: ['Технический', 'Когнитивный', 'Личностный'],
                    datasets: [{{
                        data: [technical[lastIndex], cognitive[lastIndex], personal[lastIndex]],
                        backgroundColor: ['#3B82F6', '#F59E0B', '#10B981'],
                        borderRadius: 8
                    }}]
                }},
                options: {{ responsive: true, maintainAspectRatio: true, scales: {{ y: {{ beginAtZero: true, max: 100 }} }} }}
            }});
        }}

        new Chart(document.getElementById('factorsTrendChart'), {{
            type: 'line',
            data: {{
                labels: dates,
                datasets: [
                    {{ label: 'Технический', data: technical, borderColor: '#3B82F6', fill: false }},
                    {{ label: 'Когнитивный', data: cognitive, borderColor: '#F59E0B', fill: false }},
                    {{ label: 'Личностный', data: personal, borderColor: '#10B981', fill: false }}
                ]
            }},
            options: {{ responsive: true, maintainAspectRatio: true }}
        }});
    </script>
</body>
</html>'''

        return html
